"""
DatabaseTableReader — 일반 DB형 테이블 읽기
크로스테이블이 아닌 일반 행-열 구조 엑셀용.
헤더가 상단에 있고 아래로 데이터가 쌓이는 형태.
"""
import json
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter
from langchain_core.tools import tool
from pydantic import BaseModel, Field


def _fill_merged_cells(ws) -> dict:
    merged = {}
    for r in ws.merged_cells.ranges:
        first = ws.cell(r.min_row, r.min_col).value
        for row in range(r.min_row, r.max_row + 1):
            for col in range(r.min_col, r.max_col + 1):
                merged[(row, col)] = first
    return merged


def _extract_column_names(ws, merged_values: dict, header_rows: list, data_start_col: int) -> list:
    """멀티헤더 지원: 헤더 행들의 값을 '_'로 합쳐 컬럼명 생성."""
    if not header_rows:
        return []

    max_col = ws.max_column
    col_names = []

    for col_idx in range(data_start_col, max_col + 1):
        parts = []
        for row_num in header_rows:
            val = merged_values.get((row_num, col_idx))
            if val is None:
                val = ws.cell(row_num, col_idx).value
            if val is not None and str(val).strip():
                s = str(val).strip()
                if s not in parts:
                    parts.append(s)
        col_names.append("_".join(parts) if parts else f"col_{col_idx}")

    return col_names


class DatabaseReaderInput(BaseModel):
    excel_structure: str = Field(..., description="excel_structure_parser 결과 JSON")
    header_rows: list = Field(..., description="실제 헤더 행 번호 리스트 (엑셀 1-based). 예: [3]")
    data_start_row: int = Field(..., description="데이터 시작 행 번호 (엑셀 1-based). 예: 4")
    data_start_col: int = Field(1, description="데이터 시작 열 번호 (엑셀 1-based). 기본 1.")
    sheet_name: Optional[str] = Field(None, description="시트명. None이면 첫 번째 시트 사용.")
    max_rows: Optional[int] = Field(None, description="읽을 최대 데이터 행 수. None이면 전체.")


@tool(args_schema=DatabaseReaderInput)
def database_table_reader(
    excel_structure: str,
    header_rows: list,
    data_start_row: int,
    data_start_col: int = 1,
    sheet_name: Optional[str] = None,
    max_rows: Optional[int] = None,
) -> str:
    """
    [기능]
    일반 DB형(행-열) 구조의 엑셀 데이터를 읽어 레코드 리스트로 반환합니다.
    크로스테이블이 아닌 일반 보고서 형태에 사용합니다.
    source_cell 정보로 원본 위치 역추적 가능.

    [선행 조건]
    excel_structure_parser 실행 후 LLM이 row_index를 분석하여
    header_rows, data_start_row를 직접 결정해서 전달해야 함.

    [사용 시점]
    table_type이 database일 때.
    헤더가 상단에 있고 아래로 데이터가 쌓이는 일반 표.

    [반환값]
    JSON {status, data: {rows, columns, total_rows, sheet_name}}
    각 row는 {컬럼명: 값, ..., _source_row: 행번호} 형태.
    """
    try:
        struct = json.loads(excel_structure)
        if struct.get("status") != "success":
            return json.dumps({
                "status": "error",
                "error_code": "MISSING_PREREQUISITE",
                "root_cause": "missing_prerequisite",
                "early_stop": False,
                "message": "유효한 excel_structure가 필요합니다.",
                "suggested_fix": "excel_structure_parser를 먼저 실행하세요.",
            }, ensure_ascii=False)

        file_path = struct["data"]["file_path"]
        sheets = struct["data"]["sheets"]

        # 시트 선택
        if sheet_name:
            sheet_info = next((s for s in sheets if s["name"] == sheet_name), sheets[0])
        else:
            sheet_info = sheets[0]
        target_sheet = sheet_info["name"]

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[target_sheet]
        merged_values = _fill_merged_cells(ws)

        # 코멘트 수집
        comments = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments[(cell.row, cell.column)] = cell.comment.text

        # 컬럼명 추출
        column_names = _extract_column_names(ws, merged_values, header_rows, data_start_col)

        # 데이터 읽기
        rows = []
        end_row = ws.max_row
        if max_rows:
            end_row = min(ws.max_row, data_start_row + max_rows - 1)

        for excel_row in range(data_start_row, end_row + 1):
            row_data = {"_source_row": excel_row}
            all_none = True

            for col_offset, col_name in enumerate(column_names):
                col_idx = data_start_col + col_offset
                val = merged_values.get((excel_row, col_idx))
                if val is None:
                    val = ws.cell(excel_row, col_idx).value

                if val is not None:
                    all_none = False

                col_letter = get_column_letter(col_idx)
                row_data[col_name] = val
                row_data[f"_src_{col_name}"] = f"{target_sheet}!{col_letter}{excel_row}"

                comment = comments.get((excel_row, col_idx))
                if comment:
                    row_data[f"_comment_{col_name}"] = comment

            # 완전히 빈 행 스킵
            if all_none:
                continue

            rows.append(row_data)

        wb.close()

        return json.dumps({
            "status": "success",
            "data": {
                "rows": rows,
                "columns": column_names,
                "total_rows": len(rows),
                "sheet_name": target_sheet,
                "header_rows": header_rows,
                "data_start_row": data_start_row,
            }
        }, ensure_ascii=False, default=str)

    except Exception as e:
        return json.dumps({
            "status": "error",
            "error_code": "TOOL_EXECUTION_ERROR",
            "root_cause": "unexpected_error",
            "early_stop": False,
            "message": str(e),
            "suggested_fix": "파일 경로와 파라미터를 확인하세요.",
        }, ensure_ascii=False)
