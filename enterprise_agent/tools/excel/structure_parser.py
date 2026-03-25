"""
ExcelStructureParser v2

철학:
- 파서는 판단하지 않는다. 그냥 있는 그대로 읽는다.
- 판단은 LLM이 한다.
- 파서 역할: 전체 시트를 행 번호와 함께 구조화된 JSON으로 변환

토큰 절약:
- 전체 셀을 다 주지 않고
- 각 행의 "행번호 + 첫 유의미한 값 + 전체 값 요약" 으로 압축
- LLM이 필요한 범위를 판단하면 그 범위만 다시 정밀 읽기
"""
import json
import os
from pathlib import Path
from typing import Optional

from langchain_core.tools import tool
from pydantic import BaseModel, Field


def _open_workbook(file_path: str):
    import openpyxl

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

    with open(file_path, "rb") as f:
        header = f.read(8)

    OLE2_MAGIC = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
    is_ole2 = header[:8] == OLE2_MAGIC

    if not is_ole2:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            return wb, "xlsx"
        except Exception:
            pass

    try:
        import xlrd
        wb_xls = xlrd.open_workbook(file_path)
        wb = _xlrd_to_openpyxl(wb_xls)
        return wb, "xls"
    except Exception as e:
        if "encrypted" in str(e).lower() or is_ole2:
            raise PermissionError(
                "DRM/IRM 보호 파일입니다. Excel에서 보호 해제 후 .xlsx로 저장해주세요."
            )
        raise


def _xlrd_to_openpyxl(wb_xls):
    import openpyxl
    import xlrd
    from datetime import datetime

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_idx in range(wb_xls.nsheets):
        ws_xls = wb_xls.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=ws_xls.name)
        for row in range(ws_xls.nrows):
            for col in range(ws_xls.ncols):
                cell = ws_xls.cell(row, col)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = datetime(*xlrd.xldate_as_tuple(value, wb_xls.datemode))
                    except Exception:
                        pass
                ws.cell(row=row+1, column=col+1, value=value)
    return wb


def _fill_merged_cells(ws) -> dict:
    merged = {}
    for r in ws.merged_cells.ranges:
        first = ws.cell(r.min_row, r.min_col).value
        for row in range(r.min_row, r.max_row + 1):
            for col in range(r.min_col, r.max_col + 1):
                merged[(row, col)] = first
    return merged


def _is_non_empty_cell(value) -> bool:
    return value is not None and str(value).strip() != ""


def _get_row_values(ws, merged_values: dict, excel_row: int) -> list:
    row_values = []
    for excel_col in range(1, ws.max_column + 1):
        val = merged_values.get((excel_row, excel_col))
        if val is None:
            val = ws.cell(excel_row, excel_col).value
        row_values.append(val)
    return row_values


def _build_data_sample(ws, merged_values: dict, max_rows: int = 10) -> list:
    data_sample = []
    for excel_row in range(1, min(ws.max_row, max_rows) + 1):
        data_sample.append(_get_row_values(ws, merged_values, excel_row))
    return data_sample


def _find_data_start(ws, merged_values: dict) -> dict:
    for excel_row in range(1, ws.max_row + 1):
        row_values = _get_row_values(ws, merged_values, excel_row)
        for excel_col, value in enumerate(row_values, start=1):
            if _is_non_empty_cell(value):
                return {"row": excel_row, "col": excel_col}
    return {"row": 1, "col": 1}


def _summarize_row(row_values: list) -> dict:
    """
    한 행 요약.
    LLM이 이걸 보고 "이 행이 헤더인지, 데이터인지, 메타인지" 판단함.
    """
    non_null = [
        (i, v) for i, v in enumerate(row_values)
        if v is not None and str(v).strip() != ""
    ]

    if not non_null:
        return {
            "non_null_count": 0,
            "first_value": None,
            "first_col": None,
            "value_type": "empty",
            "preview": [],
        }

    first_col, first_val = non_null[0]
    values = [v for _, v in non_null]

    num_count = sum(1 for v in values if isinstance(v, (int, float)))
    txt_count = sum(1 for v in values if isinstance(v, str))

    if num_count == len(values):
        vtype = "number"
    elif txt_count == len(values):
        vtype = "text"
    else:
        vtype = "mixed"

    return {
        "non_null_count": len(non_null),
        "first_value": str(first_val)[:50],
        "first_col": first_col + 1,
        "value_type": vtype,
        "preview": [str(v)[:20] for _, v in non_null[:5]],
    }


# ── ExcelStructureParser ─────────────────────────────────
class ExcelStructureInput(BaseModel):
    file_path: str = Field(
        ...,
        description="분석할 엑셀 파일 경로. 예: './data/sample.xlsx'"
    )


@tool(args_schema=ExcelStructureInput)
def excel_structure_parser(file_path: str) -> str:
    """
    [기능]
    엑셀 전체를 판단 없이 raw하게 읽어서 행별 요약 JSON 반환.
    LLM이 이 결과를 보고 구조를 판단하고 필요한 범위를 결정함.

    [선행 조건]
    없음. 모든 엑셀 작업의 첫 번째 단계.

    [사용 시점]
    엑셀 파일이 있는 모든 요청의 시작점.

    [반환값]
    JSON {status, data: {sheets: [{name, row_index, data_sample, data_start, merged_cells, comments}]}}
    row_index: 각 행의 행번호 + 값 요약 (LLM이 구조 판단용으로 사용)
    data_sample: 시트 상단 최대 10행의 전체 셀 값
    data_start: 시트에서 첫 번째 비어있지 않은 셀의 좌표
    """
    if ".." in file_path:
        return json.dumps({
            "status": "error",
            "error_code": "PATH_TRAVERSAL",
            "root_cause": "security_violation",
            "early_stop": True,
            "message": "허용되지 않은 경로입니다.",
        }, ensure_ascii=False)

    try:
        wb, file_format = _open_workbook(file_path)
        sheets_info = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.sheet_state == "hidden":
                continue

            merged_values = _fill_merged_cells(ws)

            comments = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment:
                        coord = f"{cell.column_letter}{cell.row}"
                        comments[coord] = cell.comment.text

            hidden_rows = [r for r, rd in ws.row_dimensions.items() if rd.hidden]
            data_sample = _build_data_sample(ws, merged_values, max_rows=10)
            data_start = _find_data_start(ws, merged_values)

            # 핵심: 전체 행 스캔 → 행별 요약
            row_index = []
            for excel_row in range(1, ws.max_row + 1):
                row_values = _get_row_values(ws, merged_values, excel_row)
                summary = _summarize_row(row_values)
                if summary["non_null_count"] > 0:
                    row_index.append({"row": excel_row, **summary})

            sheets_info.append({
                "name": sheet_name,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "row_index": row_index,
                "data_sample": data_sample,
                "data_start": data_start,
                "merged_cells": [str(r) for r in ws.merged_cells.ranges],
                "comments": comments,
                "hidden_rows": hidden_rows[:20],
            })

        wb.close()

        return json.dumps({
            "status": "success",
            "data": {
                "file_path": file_path,
                "file_format": file_format,
                "sheet_count": len(sheets_info),
                "sheets": sheets_info,
            }
        }, ensure_ascii=False, default=str)

    except FileNotFoundError as e:
        return json.dumps({
            "status": "error", "error_code": "FILE_NOT_FOUND",
            "root_cause": "invalid_path", "early_stop": False,
            "message": str(e), "suggested_fix": "파일 경로를 확인하세요.",
        }, ensure_ascii=False)

    except PermissionError as e:
        return json.dumps({
            "status": "error", "error_code": "FILE_NOT_READABLE",
            "root_cause": "drm_protected", "early_stop": True,
            "message": str(e),
        }, ensure_ascii=False)

    except Exception as e:
        return json.dumps({
            "status": "error", "error_code": "TOOL_EXECUTION_ERROR",
            "root_cause": "unexpected_error", "early_stop": False,
            "message": str(e),
            "suggested_fix": "파일이 손상되지 않았는지 확인하세요.",
        }, ensure_ascii=False)


# ── ExcelRangeReader ─────────────────────────────────────
class ExcelRangeReaderInput(BaseModel):
    file_path: str = Field(..., description="엑셀 파일 경로")
    sheet_name: str = Field(..., description="시트 이름")
    start_row: int = Field(..., description="읽기 시작 행 번호")
    end_row: int = Field(..., description="읽기 끝 행 번호")
    start_col: int = Field(1, description="읽기 시작 열 번호. 기본 1.")
    end_col: Optional[int] = Field(None, description="읽기 끝 열 번호. None이면 끝까지.")


@tool(args_schema=ExcelRangeReaderInput)
def excel_range_reader(
    file_path: str,
    sheet_name: str,
    start_row: int,
    end_row: int,
    start_col: int = 1,
    end_col: Optional[int] = None,
) -> str:
    """
    [기능]
    엑셀의 특정 범위만 정밀하게 읽어서 반환.
    LLM이 구조 파악 후 필요한 범위만 골라서 읽을 때 사용.

    [선행 조건]
    excel_structure_parser 실행 후 LLM이 범위를 결정한 다음 호출.

    [사용 시점]
    "CPK 값 읽어줘" → LLM이 CPK 행 번호 파악 → 이 툴로 그 범위만 읽기
    "측정값 분석해줘" → LLM이 측정값 시작 행 파악 → 이 툴로 읽기

    [반환값]
    JSON {status, data: {flat_table, headers, raw_rows, row_count}}
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        merged_values = _fill_merged_cells(ws)

        comments = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments[(cell.row, cell.column)] = cell.comment.text

        actual_end_col = end_col or ws.max_column
        actual_end_row = min(end_row, ws.max_row)

        raw_rows = []
        for excel_row in range(start_row, actual_end_row + 1):
            row_data = {}
            for excel_col in range(start_col, actual_end_col + 1):
                val = merged_values.get((excel_row, excel_col))
                if val is None:
                    val = ws.cell(excel_row, excel_col).value
                col_letter = ws.cell(excel_row, excel_col).column_letter
                row_data[col_letter] = {
                    "value": val,
                    "source_cell": f"{sheet_name}!{col_letter}{excel_row}",
                    "comment": comments.get((excel_row, excel_col)),
                }
            raw_rows.append({"row": excel_row, "cells": row_data})

        # flat_table: 첫 행을 헤더로 사용
        flat_table = []
        headers = {}
        if len(raw_rows) >= 2:
            header_row = raw_rows[0]["cells"]
            headers = {k: v["value"] for k, v in header_row.items()}
            for row in raw_rows[1:]:
                flat_row = {"_row": row["row"]}
                for col_letter, cell_data in row["cells"].items():
                    col_name = str(headers.get(col_letter, col_letter))
                    flat_row[col_name] = cell_data["value"]
                    flat_row[f"_src_{col_name}"] = cell_data["source_cell"]
                    if cell_data["comment"]:
                        flat_row[f"_comment_{col_name}"] = cell_data["comment"]
                flat_table.append(flat_row)

        wb.close()

        return json.dumps({
            "status": "success",
            "data": {
                "sheet_name": sheet_name,
                "range": f"{start_row}~{actual_end_row}",
                "row_count": len(raw_rows),
                "headers": headers,
                "flat_table": flat_table,
                "raw_rows": raw_rows,
            }
        }, ensure_ascii=False, default=str)

    except Exception as e:
        return json.dumps({
            "status": "error", "error_code": "TOOL_EXECUTION_ERROR",
            "root_cause": "unexpected_error", "early_stop": False,
            "message": str(e),
        }, ensure_ascii=False)
