"""
CrossTableFlattener — 실제 구현
크로스테이블 → DB형(행-열 정규화) 변환.
source_cell 추적 포함.
"""
import json
from typing import Optional
from langchain_core.tools import tool
from pydantic import BaseModel, Field
import openpyxl


def _extract_column_names(ws, merged_values: dict, header_rows: list, data_start_col: int) -> list:
    """
    멀티헤더 행들을 읽어서 컬럼명 생성.
    header_rows: 실제 엑셀 행 번호 리스트 (1-based).
    단일 헤더면 그 행의 값, 멀티헤더면 행별 값을 '_'로 합침.
    """
    if not header_rows:
        return []

    max_col = ws.max_column
    col_names = []

    for col_idx in range(data_start_col, max_col + 1):
        parts = []
        for row_num in header_rows:
            val = merged_values.get((row_num, col_idx))
            if val is None:
                val = ws.cell(row_num, col_idx).value
            if val is not None and str(val).strip():
                s = str(val).strip()
                if s not in parts:
                    parts.append(s)
        col_names.append("_".join(parts) if parts else f"col_{col_idx}")

    return col_names


def _fill_merged_cells(ws) -> dict:
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        first_val = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = first_val
    return merged_values


class CrossTableInput(BaseModel):
    excel_structure: str = Field(..., description="excel_structure_parser 결과 JSON")
    header_rows: list = Field(..., description="실제 헤더 행 번호 리스트 (엑셀 1-based). 예: [4, 5]")
    data_start_row: int = Field(..., description="데이터 시작 행 번호 (엑셀 1-based). 예: 6")
    data_start_col: int = Field(1, description="데이터 시작 열 번호 (엑셀 1-based). 기본 1.")
    sheet_name: Optional[str] = Field(None, description="시트명. None이면 첫 번째 시트 사용.")
    id_col_count: int = Field(1, description="왼쪽에서 ID로 쓸 열 개수. 예: 라인+설비=2")


@tool(args_schema=CrossTableInput)
def crosstable_flattener(
    excel_structure: str,
    header_rows: list,
    data_start_row: int,
    data_start_col: int = 1,
    sheet_name: Optional[str] = None,
    id_col_count: int = 1,
) -> str:
    """
    [기능]
    크로스테이블 형태의 엑셀을 DB형(행-열 정규화)으로 변환합니다.
    source_cell 정보를 유지하여 원본 위치 역추적 가능.

    [선행 조건]
    excel_structure_parser 실행 후 LLM이 row_index를 분석하여
    header_rows, data_start_row, id_col_count를 직접 결정해서 전달해야 함.

    [사용 시점]
    table_type이 crosstable일 때. 헤더가 날짜/항목 시퀀스인 형태.

    [반환값]
    JSON {status, data: {rows, columns, total_rows, id_columns, value_column}}
    각 row에 source_cell 포함.
    """
    try:
        struct = json.loads(excel_structure)

        if struct.get("status") != "success":
            return json.dumps({
                "status": "error",
                "error_code": "MISSING_PREREQUISITE",
                "root_cause": "missing_prerequisite",
                "early_stop": False,
                "message": "유효한 excel_structure가 필요합니다.",
                "suggested_fix": "excel_structure_parser를 먼저 실행하세요.",
            }, ensure_ascii=False)

        file_path = struct["data"]["file_path"]
        sheets = struct["data"]["sheets"]

        # 시트 선택
        if sheet_name:
            sheet_info = next((s for s in sheets if s["name"] == sheet_name), sheets[0])
        else:
            sheet_info = sheets[0]
        target_sheet = sheet_info["name"]

        # 파일 다시 열기 (원본 데이터 읽기)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[target_sheet]
        merged_values = _fill_merged_cells(ws)

        # 코멘트 수집
        comments = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments[(cell.row, cell.column)] = cell.comment.text

        # 헤더 행들에서 컬럼명 추출 (멀티헤더 지원)
        column_names = _extract_column_names(ws, merged_values, header_rows, data_start_col)

        # 헤더에서 ID 컬럼명과 값 컬럼명(날짜/항목) 분리
        id_col_names = column_names[:id_col_count]
        value_col_names = column_names[id_col_count:]

        # 데이터 읽기 + flatten
        rows = []
        for excel_row in range(data_start_row, ws.max_row + 1):
            # ID 값 읽기
            id_values = {}
            for col_offset, id_name in enumerate(id_col_names):
                col_idx = data_start_col + col_offset
                val = merged_values.get((excel_row, col_idx))
                if val is None:
                    val = ws.cell(excel_row, col_idx).value
                id_values[id_name] = val

            # 모든 ID가 None이면 빈 행 → 스킵
            if all(v is None for v in id_values.values()):
                continue

            # 각 값 컬럼을 별도 행으로 flatten
            for col_offset, val_col_name in enumerate(value_col_names):
                col_idx = data_start_col + id_col_count + col_offset
                value = merged_values.get((excel_row, col_idx))
                if value is None:
                    value = ws.cell(excel_row, col_idx).value

                # source_cell 추적
                cell_letter = ws.cell(excel_row, col_idx).column_letter
                source_cell = f"{target_sheet}!{cell_letter}{excel_row}"

                row_data = {
                    **id_values,
                    "항목": val_col_name,
                    "값": value,
                    "source_cell": source_cell,
                    "source_row": excel_row,
                    "source_col": col_idx,
                }

                # 해당 셀 코멘트 추가
                comment = comments.get((excel_row, col_idx))
                if comment:
                    row_data["comment"] = comment

                rows.append(row_data)

        wb.close()

        columns = id_col_names + ["항목", "값", "source_cell"]

        return json.dumps({
            "status": "success",
            "data": {
                "rows": rows,
                "columns": columns,
                "total_rows": len(rows),
                "id_columns": id_col_names,
                "value_column": "값",
                "dimension_column": "항목",
                "sheet_name": target_sheet,
            }
        }, ensure_ascii=False, default=str)

    except Exception as e:
        return json.dumps({
            "status": "error",
            "error_code": "TOOL_EXECUTION_ERROR",
            "root_cause": "unexpected_error",
            "early_stop": False,
            "message": str(e),
            "suggested_fix": "입력 파라미터를 확인하세요.",
        }, ensure_ascii=False)
