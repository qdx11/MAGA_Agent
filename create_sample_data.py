"""
데모용 샘플 엑셀 생성.
실제 사무실 엑셀처럼:
- 위에 제목/정보 와라라
- C4부터 시작하는 크로스테이블
- 병합셀
- 셀 메모
- 이상값 포함
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.comments import Comment
from pathlib import Path
import random
import copy


def create_sample_excel(output_path: str):
    wb = openpyxl.Workbook()

    # ── 시트1: 측정 데이터 (크로스테이블) ────────────────
    ws1 = wb.active
    ws1.title = "측정데이터"

    # 스타일 정의
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    sub_header_fill = PatternFill("solid", fgColor="9DC3E6")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # A1:F1 병합 — 제목
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "📊 라인별 일일 측정 데이터 현황"
    ws1["A1"].font = title_font
    ws1["A1"].fill = title_fill
    ws1["A1"].alignment = center

    # A2:H2 — 부제목 정보
    ws1.merge_cells("A2:D2")
    ws1["A2"] = "작성부서: 품질관리팀"
    ws1["A2"].font = Font(bold=True)

    ws1.merge_cells("E2:H2")
    ws1["E2"] = "기준: 측정값 ± 0.5 이내"
    ws1["E2"].font = Font(bold=True, color="C55A11")
    ws1["E2"].alignment = Alignment(horizontal="right")

    # A3 — 빈 행 (실제 사무실 엑셀에서 흔함)
    ws1["A3"] = ""

    # C4부터 크로스테이블 시작
    # 헤더: 날짜
    months = ["2024-01", "2024-02", "2024-03", "2024-04", "2024-05", "2024-06"]

    ws1.merge_cells("A4:B4")
    ws1["A4"] = "구분"
    ws1["A4"].font = header_font
    ws1["A4"].fill = header_fill
    ws1["A4"].alignment = center
    ws1["A4"].border = thin_border

    for i, month in enumerate(months):
        col = 3 + i  # C열부터
        cell = ws1.cell(row=4, column=col, value=month)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # 멀티헤더 두 번째 행 (소분류)
    ws1["A5"] = "라인"
    ws1["A5"].font = Font(bold=True)
    ws1["A5"].fill = PatternFill("solid", fgColor="D6E4F0")
    ws1["A5"].border = thin_border

    ws1["B5"] = "설비"
    ws1["B5"].font = Font(bold=True)
    ws1["B5"].fill = PatternFill("solid", fgColor="D6E4F0")
    ws1["B5"].border = thin_border

    for i in range(len(months)):
        col = 3 + i
        cell = ws1.cell(row=5, column=col, value="측정값(mm)")
        cell.font = Font(bold=True, size=8)
        cell.fill = PatternFill("solid", fgColor="D6E4F0")
        cell.alignment = center
        cell.border = thin_border

    # 데이터
    lines = [
        ("라인A", "설비-01"),
        ("라인A", "설비-02"),
        ("라인B", "설비-01"),
        ("라인B", "설비-02"),
        ("라인C", "설비-01"),
    ]

    # 라인A 병합
    ws1.merge_cells("A6:A7")
    ws1["A6"] = "라인A"
    ws1["A6"].alignment = center
    ws1["A6"].font = Font(bold=True)
    ws1["A6"].border = thin_border

    ws1.merge_cells("A8:A9")
    ws1["A8"] = "라인B"
    ws1["A8"].alignment = center
    ws1["A8"].font = Font(bold=True)
    ws1["A8"].border = thin_border

    ws1["A10"] = "라인C"
    ws1["A10"].font = Font(bold=True)
    ws1["A10"].border = thin_border

    # 측정 데이터 채우기 (이상값 포함)
    base_values = {
        ("라인A", "설비-01"): [1.23, 1.25, 1.21, 1.24, 1.22, 1.26],
        ("라인A", "설비-02"): [1.31, 1.29, 1.33, 1.30, 1.28, 1.32],
        ("라인B", "설비-01"): [2.10, 2.12, 2.08, 2.85, 2.11, 2.09],  # 4월 이상값!
        ("라인B", "설비-02"): [2.21, 2.19, 2.23, 2.20, 2.18, 2.22],
        ("라인C", "설비-01"): [1.55, 1.57, 1.53, 1.56, 1.54, 1.58],
    }

    for row_offset, (line, equip) in enumerate(lines):
        excel_row = 6 + row_offset
        ws1.cell(excel_row, 2, equip).border = thin_border

        values = base_values[(line, equip)]
        for col_offset, val in enumerate(values):
            col = 3 + col_offset
            cell = ws1.cell(excel_row, col, val)
            cell.border = thin_border
            cell.alignment = center
            cell.number_format = "0.000"

            # 이상값 빨간색 표시
            if (line == "라인B" and equip == "설비-01" and col_offset == 3):
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(color="9C0006", bold=True)
                # 메모 추가
                comment = Comment("기준값 초과! 설비 점검 필요\n담당: 홍길동\n확인일: 2024-04-15", "품질팀")
                cell.comment = comment

    # 일반 메모
    comment2 = Comment("2024-02월 설비 교체 후 측정값 안정화", "설비팀")
    ws1["C7"].comment = comment2

    # 열 너비 조정
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 12
    for i in range(len(months)):
        col_letter = openpyxl.utils.get_column_letter(3 + i)
        ws1.column_dimensions[col_letter].width = 14

    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[4].height = 20

    # ── 시트2: DB형 데이터 (일반 테이블) ─────────────────
    ws2 = wb.create_sheet("불량이력")

    headers2 = ["발생일", "라인", "설비", "불량유형", "불량수량", "총생산", "불량률(%)", "조치내용", "담당자"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    defect_data = [
        ("2024-01-05", "라인A", "설비-01", "치수불량", 3, 500, 0.6, "재조정", "김철수"),
        ("2024-01-12", "라인B", "설비-02", "외관불량", 5, 480, 1.04, "세척", "이영희"),
        ("2024-02-03", "라인A", "설비-02", "치수불량", 2, 510, 0.39, "재조정", "김철수"),
        ("2024-02-18", "라인C", "설비-01", "기능불량", 8, 490, 1.63, "부품교체", "박민준"),
        ("2024-03-07", "라인B", "설비-01", "외관불량", 4, 520, 0.77, "세척", "이영희"),
        ("2024-04-02", "라인B", "설비-01", "치수불량", 15, 500, 3.0, "설비점검", "박민준"),  # 이상!
        ("2024-04-15", "라인A", "설비-01", "외관불량", 2, 505, 0.40, "세척", "김철수"),
        ("2024-05-10", "라인C", "설비-01", "기능불량", 3, 495, 0.61, "부품교체", "박민준"),
        ("2024-06-01", "라인B", "설비-02", "치수불량", 1, 515, 0.19, "재조정", "이영희"),
    ]

    for row_idx, row_data in enumerate(defect_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row_idx, col_idx, val)
            cell.border = thin_border
            if col_idx == 7 and isinstance(val, float) and val > 1.5:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(color="9C0006", bold=True)

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    # ── 저장 ─────────────────────────────────────────────
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ 샘플 엑셀 생성: {output_path}")


def create_sample_v2(output_path: str, base_path: str):
    """버전 비교용 v2 파일 (일부 값 변경)"""
    import shutil
    shutil.copy(base_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["측정데이터"]

    # 일부 값 변경 (비교 데모용)
    ws["C6"] = 1.35  # 라인A 설비-01 1월 변경
    ws["E8"] = 2.45  # 라인B 설비-01 3월 변경
    ws["H10"] = 1.72  # 라인C 설비-01 6월 변경

    # 새 행 추가 (라인D)
    ws["A11"] = "라인D"
    ws["B11"] = "설비-01"
    for i, val in enumerate([1.80, 1.82, 1.79, 1.81, 1.83, 1.80], 3):
        ws.cell(11, i, val)

    wb.save(output_path)
    print(f"✅ 샘플 엑셀 v2 생성: {output_path}")


if __name__ == "__main__":
    base = "./data/measurement_v1.xlsx"
    v2 = "./data/measurement_v2.xlsx"
    create_sample_excel(base)
    create_sample_v2(v2, base)
